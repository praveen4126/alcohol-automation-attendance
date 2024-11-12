import serial
import time
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# Set up serial communication (ensure the correct COM port)
ser = serial.Serial('COM6', 9600, timeout=1)  # Adjust 'COM7' to your Arduino's port
time.sleep(2)  # Allow Arduino time to initialize

# Load the master workbook to map fingerprint IDs to names
try:
    master_wb = openpyxl.load_workbook('C:/Users/prave/Desktop/praveen/Rnaipl/Alcohol/alcohol.xlsx')
    master_sheet = master_wb.active
except FileNotFoundError:
    print("Error: Master Excel file not found.")
    exit(1)

# Load the attendance workbook and select the active sheet
try:
    wb = openpyxl.load_workbook('C:/Users/prave/Desktop/praveen/Rnaipl/Alcohol/attendance.xlsx')
    sheet = wb.active
except FileNotFoundError:
    print("Error: Attendance Excel file not found.")
    exit(1)

# Assign permanent column headers if they don't already exist
def set_column_headings():
    headers = ["Date", "Time", "ID", "Name", "ALC Level", "Attendance Status", "Test Status"]
    for col_num, header in enumerate(headers, start=1):
        if sheet.cell(row=1, column=col_num).value != header:
            sheet.cell(row=1, column=col_num, value=header)
    wb.save('C:/Users/prave/Desktop/praveen/Rnaipl/Alcohol/attendance.xlsx')

# Function to find the name for a given fingerprint ID from the master Excel file
def get_name_from_master(finger_id):
    for row in master_sheet.iter_rows(min_row=2, values_only=True):  # Assuming IDs start from the second row
        if str(row[0]) == finger_id:  # ID expected in column A
            return row[1]  # Name expected in column B
    return "Unknown User"

# Define fill colors for attendance status and test status
present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for "Present"
absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red for "Absent"
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")       # Light green for "OK"
ng_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")       # Light red for "NG"

# Function to log attendance in the attendance Excel file and apply color coding
def log_attendance(finger_id, name, alcohol_level):
    row = sheet.max_row + 1  # Move to the next available row
    current_time = datetime.now()  # Get the current date and time
    attendance_status = 'Present' if alcohol_level < 3500 else 'Absent'
    test_status = 'OK' if alcohol_level < 3500 else 'NG'

    # Fill in attendance details
    sheet.cell(row=row, column=1, value=current_time.strftime("%Y-%m-%d"))
    sheet.cell(row=row, column=2, value=current_time.strftime("%H:%M:%S"))
    sheet.cell(row=row, column=3, value=finger_id)
    sheet.cell(row=row, column=4, value=name)
    sheet.cell(row=row, column=5, value=alcohol_level)

    # Log attendance status and apply color based on the status
    attendance_cell = sheet.cell(row=row, column=6, value=attendance_status)
    attendance_cell.fill = present_fill if attendance_status == 'Present' else absent_fill

    # Log test status and apply color based on the test status
    test_status_cell = sheet.cell(row=row, column=7, value=test_status)
    test_status_cell.fill = ok_fill if test_status == 'OK' else ng_fill

    # Save the attendance Excel file
    try:
        wb.save('C:/Users/prave/Desktop/praveen/Rnaipl/Alcohol/attendance.xlsx')
        print(f"Attendance logged for Fingerprint ID {finger_id} ({name}) with Alcohol Level {alcohol_level}")
    except PermissionError:
        print("Error: Permission denied when saving the Excel file. Make sure it's not open elsewhere.")

# Set up column headings at the start of the script
set_column_headings()

finger_id, name, alcohol_level = None, None, None  # Variables to store data between messages

while True:
    if ser.in_waiting > 0:
        data = ser.readline().decode('utf-8').strip()
        print("Received from Arduino:", data)

        if "Fingerprint ID found:" in data:
            # Extract fingerprint ID from the string
            finger_id = data.split(":")[1].strip()
            print(f"Fingerprint ID: {finger_id}")

            # Get the corresponding name from the master Excel file
            name = get_name_from_master(finger_id)
            print(f"User Name: {name}")

        elif "ID:" in data and "ALC:" in data:
            # Extract fingerprint ID and alcohol level
            parts = data.split(",")
            finger_id = parts[0].split(":")[1].strip()
            alcohol_level = int(parts[1].split(":")[1].strip())
            print(f"Fingerprint ID: {finger_id}, Alcohol Level: {alcohol_level}")

        elif "Status: Present, OK" in data:
            # Log attendance for present and OK status
            if finger_id and name and alcohol_level is not None:
                log_attendance(finger_id, name, alcohol_level)
            else:
                print("Error: Missing data for logging attendance.")

        elif "Status: Absent" in data:
            # Log attendance for absent status
            if finger_id and name and alcohol_level is not None:
                log_attendance(finger_id, name, alcohol_level)
            else:
                print("Error: Missing data for logging attendance.")

    time.sleep(1)  # Delay to avoid flooding
